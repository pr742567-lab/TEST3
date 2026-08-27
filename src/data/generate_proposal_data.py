import os
import json
import re
import openpyxl

source_dir = r"C:\Users\jjpqr\Desktop\TEST2\data\개선제안\제안 실적"
output_file = r"c:\Users\jjpqr\Desktop\TEST3\src\data\proposalHistoryData.json"

proposals = []

files = sorted([f for f in os.listdir(source_dir) if f.endswith('.xlsx')])

for file_name in files:
    file_path = os.path.join(source_dir, file_name)
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # 유효한 시트 탐색
        target_sheet = None
        header_row_idx = None
        col_map = {}
        
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for r in range(1, 25):
                row_vals = [str(cell.value or '').strip().replace(' ', '') for cell in sheet[r]]
                if any('제안명' in v for v in row_vals) and any('접수번호' in v or '제안번호' in v or '성명' in v or '부서' in v for v in row_vals):
                    target_sheet = sheet
                    header_row_idx = r
                    for c_idx, val in enumerate(row_vals):
                        if '접수번호' in val or '제안번호' in val:
                            col_map['no'] = c_idx
                        elif '부서' in val:
                            col_map['dept'] = c_idx
                        elif '성명' in val or '제안자' in val:
                            col_map['name'] = c_idx
                        elif '제안명' in val:
                            col_map['title'] = c_idx
                        elif '확정' in val or '등급' in val or '1차등급' in val:
                            if 'grade' not in col_map or '확정' in val:
                                col_map['grade'] = c_idx
                        elif '제안구분' in val or '구분' in val:
                            col_map['category'] = c_idx
                    break
            if target_sheet and 'title' in col_map:
                break
        
        if not target_sheet or not header_row_idx or 'title' not in col_map:
            print(f"헤더 인식 실패: {file_name}")
            continue

        year_match = re.search(r'(\d{4})', file_name)
        default_year = year_match.group(1) if year_match else '기타'

        rows = list(target_sheet.iter_rows(min_row=header_row_idx + 1, values_only=True))
        count = 0
        for row in rows:
            if not row or not any(row):
                continue
            
            title_val = str(row[col_map['title']] or '').strip() if 'title' in col_map and col_map['title'] < len(row) else ''
            if not title_val or title_val.lower() == 'none' or title_val == '-':
                continue
                
            no_val = str(row[col_map.get('no', 0)] or '').strip() if 'no' in col_map and col_map['no'] < len(row) else ''
            dept_val = str(row[col_map.get('dept', 1)] or '').strip() if 'dept' in col_map and col_map['dept'] < len(row) else ''
            name_val = str(row[col_map.get('name', 2)] or '').strip() if 'name' in col_map and col_map['name'] < len(row) else ''
            grade_val = str(row[col_map.get('grade', 4)] or '').strip() if 'grade' in col_map and col_map['grade'] < len(row) else ''
            category_val = str(row[col_map.get('category', 6)] or '').strip() if 'category' in col_map and col_map['category'] < len(row) else ''

            if no_val.startswith('datetime') or 'datetime' in str(type(row[col_map.get('no', 0)])):
                no_val = ''

            y_match = re.search(r'^(20\d{2})', no_val)
            item_year = y_match.group(1) if y_match else default_year

            grade_val = grade_val.replace('None', '').strip()
            if not grade_val or grade_val == '-':
                grade_val = 'D'

            proposals.append({
                "id": len(proposals) + 1,
                "no": no_val if no_val and no_val != 'None' else f"{item_year}-{len(proposals)+1:04d}",
                "title": title_val,
                "name": name_val if name_val != 'None' else '',
                "dept": dept_val if dept_val != 'None' else '',
                "grade": grade_val,
                "category": category_val if category_val != 'None' and category_val != '-' else '직무개선',
                "year": item_year
            })
            count += 1
        
        print(f"[{file_name}] 추출 성공: {count}건")
    except Exception as e:
        print(f"[{file_name}] 오류 발생: {e}")

os.makedirs(os.path.dirname(output_file), exist_ok=True)
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(proposals, f, ensure_ascii=False, indent=2)

print(f"\n총 {len(proposals)}건의 제안 실적 데이터가 성공적으로 저장되었습니다: {output_file}")
